import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import date


def build_dataframe(pdf_file):

    df = extract_pdf_tables(pdf_file)

    df["Produit"] = (
        df["Produit"]
        .astype(str)
        .str.strip()
    )

    df["Origine"] = (
        df["Origine"]
        .astype(str)
        .str.strip()
    )

    for col in ["Prix", "Minimum", "Maximum"]:

        df[col] = clean_numeric(
            df[col]
        )

    df["Prix Mafate"] = (
        df["Prix"] * 1.30
    ).round(2)

    return (
        df
        .sort_values("Produit")
        .reset_index(drop=True)
    )

def extract_pdf_tables(pdf_file):

    rows = []

    with pdfplumber.open(pdf_file) as pdf:

        for page in pdf.pages:

            table = page.extract_table()

            if not table:
                continue

            for row in table:

                if not row:
                    continue

                if row[0] == "Produit":
                    continue

                if ((
                    row[0]
                    and "MARCHES FORAINS" in str(row[0])
                ) or (
                    row[0]
                    and "GRANDES & MOYENNES SURFACES" in str(row[0])
                )):
                    continue

                if len(row) >= 6:

                    rows.append(row[:6])

    return pd.DataFrame(
        rows,
        columns=[
            "Produit",
            "Origine",
            "Prix",
            "Variation",
            "Minimum",
            "Maximum"
        ]
    )

def build_average_dataframe(
    forains_df,
    gms_df
):

    merged = (
        forains_df[
            ["Produit", "Prix"]
        ]
        .rename(
            columns={
                "Prix": "Prix Forains"
            }
        )
        .merge(
            gms_df[
                ["Produit", "Prix"]
            ].rename(
                columns={
                    "Prix": "Prix GMS"
                }
            ),
            on="Produit",
            how="outer"
        )
    )

    merged["Prix Moyen"] = (
        merged[
            ["Prix Forains", "Prix GMS"]
        ]
        .mean(axis=1)
        .round(2)
    )

    merged["Prix Mafate"] = (
        merged["Prix Moyen"] * 1.30
    ).round(2)

    return (
        merged
        .sort_values("Produit")
        .reset_index(drop=True)
    )

def clean_numeric(series):

    return pd.to_numeric(
        series
            .astype(str)
            .str.replace(",", ".", regex=False),
        errors="coerce"
    )

def export_excel(
    forains_df,
    gms_df,
    average_df,
    output_file
):

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        forains_df.to_excel(
            writer,
            sheet_name="Marches Forains",
            index=False
        )

        gms_df.to_excel(
            writer,
            sheet_name="GMS",
            index=False
        )

        average_df.to_excel(
            writer,
            sheet_name="Moyenne",
            index=False
        )

    wb = load_workbook(output_file)

    for sheet in wb.sheetnames:

        ws = wb[sheet]

        ws.auto_filter.ref = ws.dimensions

        for column in ws.columns:

            max_length = 0

            letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

                except:

                    pass

            ws.column_dimensions[
                letter
            ].width = max_length + 2

    wb.save(output_file)
